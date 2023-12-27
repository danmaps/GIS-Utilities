import openpyxl
import sys
from openpyxl.utils import get_column_letter, column_index_from_string

def identify_column(sheet, column_identifier):
    """
    Identify the column index based on the identifier which can be
    a letter, a number, or a heading.
    """
    if column_identifier.isnumeric():
        return int(column_identifier)
    elif column_identifier.isalpha():
        return column_index_from_string(column_identifier.upper())
    else:
        # Assuming it's a heading; find the column with this heading
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=i).value == column_identifier:
                return i
        raise ValueError(f"Column heading '{column_identifier}' not found.")

def main(file_path, sheet_name, *columns):
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} does not exist.")
        return

    wb = openpyxl.load_workbook(file_path)
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
        if not sheet_name:
            print(f"Warning: No sheet name specified. Using the first sheet '{sheet.title}' for summarization.")

    # Check for existing Summary sheet and create a unique name
    summary_sheet_title = "Summary"
    count = 1
    while summary_sheet_title in wb.sheetnames:
        summary_sheet_title = f"Summary_{count}"
        count += 1
    summary_sheet = wb.create_sheet(summary_sheet_title)

    # Write headers to the summary sheet
    headers = ["Column Heading", "Count", "Total", "Percentage"]
    summary_sheet.append(headers)

    total_rows_formula = f"=COUNTA({sheet.title}!A:A)-1"

    for col in columns:
        col_index = identify_column(sheet, col)
        col_letter = get_column_letter(col_index)
        heading = sheet.cell(row=1, column=col_index).value

        count_formula = f"=COUNTA({sheet.title}!{col_letter}:{col_letter})-1"
        percentage_formula = f"={get_column_letter(2)}{summary_sheet.max_row+1}/{get_column_letter(3)}{summary_sheet.max_row+1}"

        summary_sheet.append([heading, count_formula, total_rows_formula, percentage_formula])

    # Save the workbook
    wb.save(file_path)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python script.py <file_path> <sheet_name> <column1> <column2> ...")
        sys.exit(1)

    main(sys.argv[1], sys.argv[2], *sys.argv[3:])
